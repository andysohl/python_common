# -*- coding: utf-8 -*-
import io
from typing import Any

import msoffcrypto
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException


def open_excel(source: Any, password: str = None) -> Workbook:
    """
    Open an Excel file and return an openpyxl Workbook. Cached values will be in the workbook.
    If encrypted, only file sources can be used.

    :param source: The file/network/memory location to load from
    :param password: The password is encrypted.
    :return: The opened Workbook
    """

    if password is None:
        if type(source) == bytes:
            source = io.BytesIO(source)
        return load_workbook(filename=source, read_only=False, keep_vba=True, data_only=True)
    else:
        decrypted_wb = io.BytesIO()
        try:
            with open(source, "rb") as f:
                file = msoffcrypto.OfficeFile(f)
                file.load_key(password=password)
                file.decrypt(decrypted_wb)

            # Create a handle on the workbook for navigation. Needs RW since some features don't work otherwise.
            return load_workbook(filename=decrypted_wb, read_only=False, data_only=True)
        except (InvalidFileException, FileNotFoundError, msoffcrypto.exceptions.InvalidKeyError) as err:
            # FIXME: logging
            raise err
